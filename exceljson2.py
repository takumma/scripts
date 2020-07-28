import openpyxl
import json
import codecs

book = openpyxl.load_workbook('C:/Users/taku3/work/scripts/1365344_1-0207r.xlsx')
sheets = book['07 果実類']
path_w = 'C:/Users/taku3/work/scripts/fruit2.json'
fruits_list = [{
        "food_id": 0,
        "name": ""
        }]
for i in range(2, 175):
    food_id = sheets.cell(row = i, column = 2).value
    name = sheets.cell(row = i, column = 4).value
    dietary_fiber = sheets.cell(row = i, column = 21).value
    potassium = sheets.cell(row = i, column = 24).value
    iron = sheets.cell(row = i, column = 28).value
    vitamin_b1 = sheets.cell(row = i, column = 48).value
    vitamin_c = sheets.cell(row = i, column = 56).value
    
    if dietary_fiber == 'Tr':
        dietary_fiber = 0
    else:
        dietary_fiber = float(dietary_fiber)
    if potassium == 'Tr':
        potassium = 0
    else:
        potassium = float(potassium)
    if iron == 'Tr':
        iron = 0
    else:
        iron = float(iron)
    if vitamin_b1 == 'Tr':
        vitamin_b1 = 0
    else:
        vitamin_b1 = float(vitamin_b1)
    if vitamin_c == 'Tr':
        vitamin_c = 0
    else:
        vitamin_c = float(vitamin_c)

    
    if name.split("　")[0][0] == '(' or name.split("　")[0][0] == '（':
        name = name.split("　")[1]
    else:
        name = name.split("　")[0]
    
    if fruits_list[-1]['name'] != name:
        fruits_list.append({
        "food_id": int(food_id),
        "name": name,
        "dietary_fiber": float(dietary_fiber),
        "potassium": float(potassium),
        "iron": float(iron),
        "vitamin_b1": float(vitamin_b1),
        "vitamin_c": float(vitamin_c),
        })
    
data_dict = {
    "data": "fruits",
    "fruits": fruits_list
}
with open(path_w, mode = 'a', encoding = 'utf-8') as f:
    f.write(json.dumps(data_dict, ensure_ascii = False, indent = 4))
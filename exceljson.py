import openpyxl
import json
import codecs

book = openpyxl.load_workbook('C:/Users/taku3/Downloads/1365344_1-0207r.xlsx')
sheets = book['07 果実類']
path_w = 'C:/Users/taku3/work/scripts/fruit.json'

for i in range(2,175):
	id = sheets.cell(row=i, column=2).value
	name = sheets.cell(row=i, column=4).value
	fruits = {
	                    "id": id,
	                    "name": name,
             }
	with open(path_w, mode='a', encoding='utf-8') as f:
		f.write(json.dumps(fruits, ensure_ascii=False,indent=4))
import openpyxl
import json

load_book = openpyxl.load_workbook('C:/Users/taku3/work/scripts/1365344_1-0206r11.xlsx')
sheet = load_book['06 野菜類']
json_path = 'C:/Users/taku3/work/sumuppi/json/vegetable.json'

fruits_list = [{
    "food_id": 0,
    "name": ""
}]
for i in range(9, 371):
    food_id = sheet.cell(row = i, column = 2).value
    name = sheet.cell(row = i, column = 4).value
    dietary_fiber = sheet.cell(row = i, column = 21).value
    potassium = sheet.cell(row = i, column = 24).value
    iron = sheet.cell(row = i, column = 28).value
    vitamin_b1 = sheet.cell(row = i, column = 48).value
    vitamin_c = sheet.cell(row = i, column = 56).value

    food_id = int(food_id)

    if dietary_fiber == 'Tr':
        dietary_fiber = 0
    if potassium == 'Tr':
        potassium = 0
    if iron == 'Tr':
        iron = 0
    if vitamin_b1 == 'Tr':
        vitamin_b1 = 0
    if vitamin_c == 'Tr':
        vitamin_c = 0

    if name.split("　")[0][0] == '(' or name.split("　")[0][0] == '（':
        name = name.split("　")[1]
    else:
        name = name.split("　")[0]
    
    if fruits_list[-1]['name'] != name:
        fruits_list.append({
            "food_id": food_id,
            "name": name,
            "dietary_fiber": dietary_fiber,
            "potassium": potassium,
            "iron": iron,
            "vitamin_b1": vitamin_b1,
            "vitamin_c": vitamin_c,
        })
    
fruits_list.pop(0)

data_dict = {
    "data": "fruits",
    "fruits": fruits_list
}

with open(json_path, mode = 'w', encoding = 'utf-8') as f:
    f.write(json.dumps(data_dict, ensure_ascii = False, indent = 4))